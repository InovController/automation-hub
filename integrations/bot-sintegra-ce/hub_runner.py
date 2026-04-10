import asyncio
import json
import os
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from init_navegador import iniciar_navegador

OUTPUT_FILENAME = "consulta_sintegra_ce.xlsx"
HEADERS = [
    "CNPJ",
    "CNAE Principal(Nacional)",
    "CNAE Principal(Arrec/Fiscal)",
    "Regime de Recolhimento",
    "Empresa do Simples?",
    "Status",
]


def emit_log(level: str, message: str) -> None:
    print(f"AH_LOG|{level}|{message}", flush=True)


def emit_progress(value: int, message: str) -> None:
    print(f"AH_PROGRESS|{value}|{message}", flush=True)


def get_env(name: str, default: str | None = None) -> str:
    value = os.environ.get(name, default)
    if value is None:
        raise EnvironmentError(f"Variavel de ambiente obrigatoria ausente: {name}")
    return value


def read_json_file(path_str: str | None) -> dict:
    if not path_str:
        return {}

    path = Path(path_str)
    if not path.exists():
        return {}

    return json.loads(path.read_text(encoding="utf-8-sig"))


def load_parameters() -> dict:
    parameters = read_json_file(os.environ.get("AUTOMATION_PARAMETERS_FILE"))
    if parameters:
        return parameters

    metadata_dir = os.environ.get("AUTOMATION_METADATA_DIR")
    if metadata_dir:
        fallback = Path(metadata_dir) / "parameters.json"
        return read_json_file(str(fallback))

    return {}


def to_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "sim", "yes", "on"}


def sanitize_cnpj(value: object) -> str:
    return (
        str(value or "")
        .strip()
        .replace(".", "")
        .replace("/", "")
        .replace("-", "")
    )


def find_first_excel(input_dir: Path) -> Path:
    candidates = sorted(input_dir.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("Nenhum arquivo .xlsx foi enviado para a consulta do Sintegra.")
    return candidates[0]


def ensure_headers(sheet) -> None:
    current_headers = [sheet.cell(row=1, column=index).value for index in range(1, 7)]
    if current_headers == HEADERS:
        return

    sheet.insert_rows(1)
    for index, header in enumerate(HEADERS, start=1):
        sheet.cell(row=1, column=index, value=header)


def formatar_cabecalho(sheet) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 7):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def formatar_corpo(sheet) -> None:
    data_font = Font(color="000000", size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment


async def consultar_cnpj(page, cnpj: str) -> None:
    btn_dropdown = await page.find("Selecione o perfil de contribuinte", best_match=True)
    await btn_dropdown.click()
    await asyncio.sleep(1.2)

    opcao_cnpj = await page.find("CNPJ", best_match=True)
    await opcao_cnpj.click()
    await asyncio.sleep(1.0)

    input_cnpj = await page.select("#numcnpjcgf")
    await input_cnpj.send_keys(cnpj)
    await asyncio.sleep(0.8)

    btn_pesquisar = await page.find("Pesquisar", best_match=True)
    await btn_pesquisar.click()
    await asyncio.sleep(3.0)


async def extrair_dados(page) -> dict:
    await asyncio.sleep(2.5)

    if await page.xpath('//h2[contains(text(), "Nenhum contribuinte encontrado")]'):
        await page.get("https://consultapublica.sefaz.ce.gov.br/sintegra/preparar-consultar")
        await asyncio.sleep(2)
        return {
            "cnae_nacional": "SEM CGF",
            "cnae_arrec": "",
            "regime": "",
            "simples": "",
            "status": "Consultado",
        }

    dados = {
        "cnae_nacional": (await page.xpath("//tr[th[text()='CNAE Principal(Nacional)']]/td"))[0].text.strip(),
        "cnae_arrec": (await page.xpath("//tr[th[text()='CNAE Principal(Arrec/Fiscal)']]/td"))[0].text.strip(),
        "regime": (await page.xpath("//tr[th[text()='Regime de Recolhimento']]/td"))[0].text.strip(),
        "simples": (await page.xpath("//tr[th[text()='Opção Simples']]/td"))[0].text.strip(),
        "status": "Consultado",
    }
    return dados


def atualizar_linha(sheet, row_index: int, dados: dict) -> None:
    sheet.cell(row=row_index, column=2, value=dados["cnae_nacional"])
    sheet.cell(row=row_index, column=3, value=dados["cnae_arrec"])
    sheet.cell(row=row_index, column=4, value=dados["regime"])
    sheet.cell(row=row_index, column=5, value=dados["simples"])
    sheet.cell(row=row_index, column=6, value=dados["status"])


def calcular_total(sheet) -> int:
    total = 0
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        cnpj = sanitize_cnpj(row[0].value)
        status = row[5].value if len(row) >= 6 else None
        if len(cnpj) >= 11 and status not in ("Consultado", "Erro", "Erro na extração"):
            total += 1
    return total


async def run() -> int:
    execution_id = get_env("AUTOMATION_EXECUTION_ID", "execucao-sem-id")
    input_dir = Path(get_env("AUTOMATION_INPUT_DIR"))
    output_dir = Path(get_env("AUTOMATION_OUTPUT_DIR"))
    parameters = load_parameters()
    headless = not to_bool(parameters.get("mostrarNavegador"), default=False)

    output_dir.mkdir(parents=True, exist_ok=True)

    input_file = find_first_excel(input_dir)
    output_file = output_dir / OUTPUT_FILENAME
    shutil.copy2(input_file, output_file)

    emit_progress(5, "Preparando planilha de trabalho")
    emit_log("info", f"Execucao recebida: {execution_id}")
    emit_log("info", f"Arquivo de entrada: {input_file.name}")
    emit_log("info", f"Arquivo de saida: {output_file.name}")

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active
    ensure_headers(sheet)
    workbook.save(output_file)

    total = calcular_total(sheet)
    if total == 0:
        formatar_cabecalho(sheet)
        formatar_corpo(sheet)
        workbook.save(output_file)
        emit_log("warn", "Nenhum CNPJ pendente encontrado para processar.")
        emit_progress(100, "Execucao concluida sem itens pendentes")
        return 0

    page = None
    processados = 0

    try:
        emit_progress(10, "Abrindo navegador e acessando portal do Sintegra")
        page = await iniciar_navegador(headless=headless)
        emit_log("info", "Portal do Sintegra carregado com sucesso.")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            cnpj_cell = row[0]
            status_cell = row[5] if len(row) >= 6 else None
            cnpj = sanitize_cnpj(cnpj_cell.value)

            if len(cnpj) < 11:
                continue

            if status_cell and status_cell.value in ("Consultado", "Erro", "Erro na extração"):
                continue

            emit_log("info", f"Consultando CNPJ {cnpj} na linha {cnpj_cell.row}.")

            try:
                await consultar_cnpj(page, cnpj)
                dados = await extrair_dados(page)
                atualizar_linha(sheet, cnpj_cell.row, dados)
                workbook.save(output_file)
                processados += 1
                progresso = 10 + int(round((processados / total) * 85))
                emit_progress(progresso, f"Processando CNPJs ({processados}/{total})")
            except Exception as exc:
                atualizar_linha(
                    sheet,
                    cnpj_cell.row,
                    {
                        "cnae_nacional": "",
                        "cnae_arrec": "",
                        "regime": "",
                        "simples": "",
                        "status": "Erro",
                    },
                )
                workbook.save(output_file)
                emit_log("error", f"Linha {cnpj_cell.row} ({cnpj}): {exc}")

        formatar_cabecalho(sheet)
        formatar_corpo(sheet)
        workbook.save(output_file)
        emit_log("info", "Planilha final gravada com sucesso.")
        emit_progress(100, "Consulta Sintegra concluida")
        return 0
    finally:
        if page and hasattr(page, "browser") and page.browser is not None:
            try:
                await page.browser.stop()
            except Exception as exc:
                emit_log("warn", f"Nao foi possivel fechar o navegador corretamente: {exc}")


if __name__ == "__main__":
    try:
        raise SystemExit(asyncio.run(run()))
    except Exception as exc:
        emit_log("error", str(exc))
        raise SystemExit(1)
