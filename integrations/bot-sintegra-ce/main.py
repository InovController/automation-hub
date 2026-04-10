import openpyxl
import asyncio
import threading
from openpyxl.styles import Font, Alignment, PatternFill

from init_navegador import iniciar_navegador


# ====================== FUNÇÕES ASSÍNCRONAS ======================
async def consulta(page, cnpj):
    try:
        # Dropdown
        btn_dropdown = await page.find("Selecione o perfil de contribuinte", best_match=True)
        await btn_dropdown.click()
        await asyncio.sleep(1.2)

        # Opção CNPJ
        opcao_cnpj = await page.find("CNPJ", best_match=True)
        await opcao_cnpj.click()
        await asyncio.sleep(1.0)

        # Campo CNPJ
        input_cnpj = await page.select("#numcnpjcgf")
        await input_cnpj.send_keys(str(cnpj).strip())
        await asyncio.sleep(0.8)

        # Botão Pesquisar
        btn_pesquisar = await page.find("Pesquisar", best_match=True)
        await btn_pesquisar.click()
        await asyncio.sleep(3.0)

    except Exception as e:
        print(f"Erro na consulta do CNPJ {cnpj}: {e}")
        raise


async def extração(page, cell, workbook, sheet, file_path):
    await asyncio.sleep(2.5)

    # Verifica mensagem de "não encontrado"
    if await page.xpath('//h2[contains(text(), "Nenhum contribuinte encontrado")]'):
        sheet.cell(row=cell.row, column=2, value="SEM CGF")
        sheet.cell(row=cell.row, column=6, value="Consultado")
        workbook.save(file_path)
        await page.get("https://consultapublica.sefaz.ce.gov.br/sintegra/preparar-consultar")
        await asyncio.sleep(2)
        return

    try:
        cnae_nacional = (await page.xpath("//tr[th[text()='CNAE Principal(Nacional)']]/td"))[0].text.strip()
        cnae_arrec    = (await page.xpath("//tr[th[text()='CNAE Principal(Arrec/Fiscal)']]/td"))[0].text.strip()
        regime        = (await page.xpath("//tr[th[text()='Regime de Recolhimento']]/td"))[0].text.strip()
        simples       = (await page.xpath("//tr[th[text()='Opção Simples']]/td"))[0].text.strip()

        sheet.cell(row=cell.row, column=2, value=cnae_nacional)
        sheet.cell(row=cell.row, column=3, value=cnae_arrec)
        sheet.cell(row=cell.row, column=4, value=regime)
        sheet.cell(row=cell.row, column=5, value=simples)
        sheet.cell(row=cell.row, column=6, value="Consultado")
        workbook.save(file_path)

    except Exception as e:
        print(f"Erro na extração da linha {cell.row}: {e}")
        sheet.cell(row=cell.row, column=6, value="Erro na extração")
        workbook.save(file_path)


def formatar_cabeçalho(sheet):
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 7):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def formatar_corpo(sheet):
    data_font = Font(color="000000", size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment


# ====================== MAIN (chamado pela interface) ======================
def main(interface, file_path):
    def run_nodriver():
        async def worker():
            page = None
            try:
                page = await iniciar_navegador()
                interface.thread.log_signal.emit("✅ Navegador aberto com sucesso!")

                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                # Cabeçalho
                headers = ["CNPJ", "CNAE Principal(Nacional)", "CNAE Principal(Arrec/Fiscal)",
                           "Regime de Recolhimento", "Empresa do Simples?", "Status"]
                sheet.insert_rows(1)
                for i, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=i, value=header)
                workbook.save(file_path)

                total = sum(1 for c in sheet["A"] if c.value and c.row > 1)
                processados = 0

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    cnpj_cell = row[0]
                    status_cell = row[5] if len(row) >= 6 else None

                    if status_cell and status_cell.value in ("Consultado", "Erro na extração"):
                        continue

                    cnpj = str(cnpj_cell.value or "").strip().replace(".", "").replace("/", "").replace("-", "")
                    if len(cnpj) < 11:
                        continue

                    try:
                        await consulta(page, cnpj)
                        await extração(page, cnpj_cell, workbook, sheet, file_path)
                        processados += 1
                        perc = int(round(processados / total * 100))
                        interface.thread.log_signal.emit(f"Progresso: {perc}% ({processados}/{total})")
                    except Exception as e:
                        interface.thread.log_signal.emit(f"❌ Erro na linha {cnpj_cell.row}: {e}")
                        sheet.cell(row=cnpj_cell.row, column=6, value="Erro")
                        workbook.save(file_path)

                formatar_cabeçalho(sheet)
                formatar_corpo(sheet)
                workbook.save(file_path)
                interface.thread.log_signal.emit("🎉 Consulta concluída com sucesso!")

            except Exception as e:
                interface.thread.log_signal.emit(f"❌ Erro fatal: {e}")
                print(f"Erro fatal: {e}")

            finally:
                # Fechamento seguro
                try:
                    if page and hasattr(page, 'browser') and page.browser is not None:
                        await page.browser.stop()
                except Exception as close_e:
                    print(f"Aviso ao fechar: {close_e}")

        asyncio.run(worker())

    # Executa em thread separada
    threading.Thread(target=run_nodriver, daemon=True).start()